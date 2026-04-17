import csv
import io
from datetime import datetime
from flask import Blueprint, request, Response, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.models import db, User, Transaction, Bill, Client, Product, Quote, Order
from functools import wraps
from flask import abort

def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = get_current_user()
        if not user or user.role != 'admin':
            abort(403)
        return fn(*args, **kwargs)
    return wrapper

import_export_bp = Blueprint('import_export', __name__)


def get_current_user():
    return User.query.get(get_jwt_identity())


def parse_date(date_str):
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        return None


def safe(obj, attr, default=''):
    """Acessa atributo com segurança — evita AttributeError se campo não existir no model."""
    val = getattr(obj, attr, default)
    return val if val is not None else default


def fmt_date(val):
    if not val:
        return ''
    try:
        return val.strftime('%d/%m/%Y')
    except Exception:
        return str(val)


def fmt_money(val):
    try:
        return f'{float(val):.2f}'
    except Exception:
        return '0.00'


def make_response_csv(rows, headers, filename):
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    writer.writerows(rows)
    output.seek(0)
    bom = '\ufeff'
    return Response(
        bom + output.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Access-Control-Expose-Headers': 'Content-Disposition',
        }
    )


def make_response_xlsx(rows, headers, filename):
    """Gera xlsx sem openpyxl — usa CSV com extensão xlsx como fallback seguro,
    ou openpyxl se disponível."""
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, '') for h in headers])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Access-Control-Expose-Headers': 'Content-Disposition',
            }
        )
    except ImportError:
        # fallback: CSV renomeado para xlsx (Excel abre normalmente)
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
        output.seek(0)
        bom = '\ufeff'
        return Response(
            bom + output.getvalue(),
            mimetype='text/csv; charset=utf-8',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Access-Control-Expose-Headers': 'Content-Disposition',
            }
        )


def build_response(rows, headers, base_filename, fmt):
    if fmt == 'xlsx':
        return make_response_xlsx(rows, headers, f'{base_filename}.xlsx')
    return make_response_csv(rows, headers, f'{base_filename}.csv')


# ─────────────────────────────────────────────
# EXPORTAÇÕES
# ─────────────────────────────────────────────

@import_export_bp.route('/export/transactions', methods=['GET'])
@jwt_required()
def export_transactions():
    user      = get_current_user()
    date_from = parse_date(request.args.get('date_from'))
    date_to   = parse_date(request.args.get('date_to'))
    fmt       = request.args.get('format', 'csv')

    q = Transaction.query.filter_by(company_id=user.company_id)
    if date_from:
        q = q.filter(Transaction.date >= date_from)
    if date_to:
        q = q.filter(Transaction.date <= date_to)

    headers = ['ID', 'Tipo', 'Descrição', 'Valor', 'Categoria', 'Data', 'Status', 'Origem', 'Observações']
    rows = []
    for t in q.order_by(Transaction.date.desc()).all():
        rows.append({
            'ID':          safe(t, 'id'),
            'Tipo':        safe(t, 'type'),
            'Descrição':   safe(t, 'description'),
            'Valor':       fmt_money(safe(t, 'amount', 0)),
            'Categoria':   safe(t, 'category'),
            'Data':        fmt_date(safe(t, 'date', None)),
            'Status':      safe(t, 'status'),
            'Origem':      safe(t, 'source', 'manual'),
            'Observações': safe(t, 'notes'),
        })

    stamp = datetime.now().strftime('%Y%m%d')
    return build_response(rows, headers, f'transacoes_{stamp}', fmt)


@import_export_bp.route('/export/bills', methods=['GET'])
@jwt_required()
def export_bills():
    user      = get_current_user()
    date_from = parse_date(request.args.get('date_from'))
    date_to   = parse_date(request.args.get('date_to'))
    fmt       = request.args.get('format', 'csv')

    q = Bill.query.filter_by(company_id=user.company_id)
    if date_from:
        q = q.filter(Bill.due_date >= date_from)
    if date_to:
        q = q.filter(Bill.due_date <= date_to)

    headers = ['ID', 'Tipo', 'Descrição', 'Valor', 'Categoria', 'Vencimento', 'Status', 'Recorrente', 'Observações']
    rows = []
    for b in q.order_by(Bill.due_date.desc()).all():
        rows.append({
            'ID':          safe(b, 'id'),
            'Tipo':        safe(b, 'type'),
            'Descrição':   safe(b, 'description'),
            'Valor':       fmt_money(safe(b, 'amount', 0)),
            'Categoria':   safe(b, 'category'),
            'Vencimento':  fmt_date(safe(b, 'due_date', None)),
            'Status':      safe(b, 'status'),
            'Recorrente':  'Sim' if safe(b, 'is_recurring', False) else 'Não',
            'Observações': safe(b, 'notes'),
        })

    stamp = datetime.now().strftime('%Y%m%d')
    return build_response(rows, headers, f'contas_{stamp}', fmt)


@import_export_bp.route('/export/clients', methods=['GET'])
@jwt_required()
def export_clients():
    user      = get_current_user()
    date_from = parse_date(request.args.get('date_from'))
    date_to   = parse_date(request.args.get('date_to'))
    fmt       = request.args.get('format', 'csv')

    q = Client.query.filter_by(company_id=user.company_id)
    if date_from:
        q = q.filter(Client.created_at >= date_from)
    if date_to:
        q = q.filter(Client.created_at <= date_to)

    headers = ['ID', 'Nome', 'Email', 'Telefone', 'Documento', 'Endereço', 'Criado em', 'Observações']
    rows = []
    for c in q.order_by(Client.name).all():
        rows.append({
            'ID':          safe(c, 'id'),
            'Nome':        safe(c, 'name'),
            'Email':       safe(c, 'email'),
            'Telefone':    safe(c, 'phone'),
            'Documento':   safe(c, 'document'),
            'Endereço':    safe(c, 'address'),
            'Criado em':   safe(c, 'created_at'),
            'Observações': safe(c, 'notes'),
        })

    stamp = datetime.now().strftime('%Y%m%d')
    return build_response(rows, headers, f'clientes_{stamp}', fmt)


@import_export_bp.route('/export/products', methods=['GET'])
@jwt_required()
def export_products():
    user = get_current_user()
    fmt  = request.args.get('format', 'csv')

    headers = ['ID', 'Nome', 'SKU', 'Tipo', 'Categoria', 'Preço de Venda', 'Custo',
               'Estoque', 'Estoque Mínimo', 'Unidade', 'Descrição', 'Ativo']
    rows = []
    for p in Product.query.filter_by(company_id=user.company_id).order_by(Product.name).all():
        rows.append({
            'ID':              safe(p, 'id'),
            'Nome':            safe(p, 'name'),
            'SKU':             safe(p, 'sku'),
            'Tipo':            safe(p, 'type'),
            'Categoria':       safe(p, 'category'),
            'Preço de Venda':  fmt_money(safe(p, 'price', 0)),
            'Custo':           fmt_money(safe(p, 'cost', 0)),
            'Estoque':         safe(p, 'stock_qty', 0),
            'Estoque Mínimo':  safe(p, 'stock_min', 0),
            'Unidade':         safe(p, 'unit'),
            'Descrição':       safe(p, 'description'),
            'Ativo':           'Sim' if safe(p, 'active', True) else 'Não',
        })

    stamp = datetime.now().strftime('%Y%m%d')
    return build_response(rows, headers, f'produtos_{stamp}', fmt)


@import_export_bp.route('/export/quotes', methods=['GET'])
@jwt_required()
def export_quotes():
    user      = get_current_user()
    date_from = parse_date(request.args.get('date_from'))
    date_to   = parse_date(request.args.get('date_to'))
    fmt       = request.args.get('format', 'csv')

    q = Quote.query.filter_by(company_id=user.company_id)
    if date_from:
        q = q.filter(Quote.created_at >= date_from)
    if date_to:
        q = q.filter(Quote.created_at <= date_to)

    headers = ['Número', 'Cliente', 'Status', 'Subtotal', 'Desconto %', 'Total', 'Validade', 'Criado em', 'Observações']
    rows = []
    for qt in q.order_by(Quote.created_at.desc()).all():
        client_obj  = getattr(qt, 'client', None)
        client_name = client_obj.name if client_obj else safe(qt, 'client_name')
        rows.append({
            'Número':      safe(qt, 'number') or safe(qt, 'id'),
            'Cliente':     client_name,
            'Status':      safe(qt, 'status'),
            'Subtotal':    fmt_money(safe(qt, 'subtotal', 0)),
            'Desconto %':  fmt_money(safe(qt, 'discount', 0)),
            'Total':       fmt_money(safe(qt, 'total', 0)),
            'Validade':    fmt_date(safe(qt, 'valid_until', None)),
            'Criado em':   fmt_date(safe(qt, 'created_at', None)),
            'Observações': safe(qt, 'notes'),
        })

    stamp = datetime.now().strftime('%Y%m%d')
    return build_response(rows, headers, f'orcamentos_{stamp}', fmt)


@import_export_bp.route('/export/sales', methods=['GET'])
@jwt_required()
def export_sales():
    user      = get_current_user()
    date_from = parse_date(request.args.get('date_from'))
    date_to   = parse_date(request.args.get('date_to'))
    fmt       = request.args.get('format', 'csv')

    q = Order.query.filter_by(company_id=user.company_id)
    if date_from:
        q = q.filter(Order.created_at >= date_from)
    if date_to:
        q = q.filter(Order.created_at <= date_to)

    headers = ['Número', 'Tipo', 'Cliente', 'Vendedor', 'Status', 'Total', 'Origem', 'Criado em', 'Concluído em', 'Observações']
    rows = []
    for o in q.order_by(Order.created_at.desc()).all():
        client_obj = getattr(o, 'client', None)
        seller_obj = getattr(o, 'seller', None)
        rows.append({
            'Número':       safe(o, 'order_number') or safe(o, 'id'),
            'Tipo':         safe(o, 'order_type'),
            'Cliente':      client_obj.name if client_obj else '',
            'Vendedor':     seller_obj.name if seller_obj else '',
            'Status':       safe(o, 'status'),
            'Total':        fmt_money(safe(o, 'total', 0)),
            'Origem':       'Orçamento' if safe(o, 'quote_id') else 'Direto',
            'Criado em':    fmt_date(safe(o, 'created_at', None)),
            'Concluído em': fmt_date(safe(o, 'completed_at', None)),
            'Observações':  safe(o, 'notes'),
        })

    stamp = datetime.now().strftime('%Y%m%d')
    return build_response(rows, headers, f'vendas_{stamp}', fmt)


# ─────────────────────────────────────────────
# TEMPLATES PARA DOWNLOAD
# ─────────────────────────────────────────────

TEMPLATES = {
    'transactions': {
        'headers': ['Tipo', 'Descrição', 'Valor', 'Categoria', 'Data', 'Status', 'Observações'],
        'example': {'Tipo': 'receita', 'Descrição': 'Venda de produto', 'Valor': '150.00',
                    'Categoria': 'Vendas', 'Data': '01/01/2025', 'Status': 'confirmado', 'Observações': ''},
    },
    'bills': {
        'headers': ['Tipo', 'Descrição', 'Valor', 'Categoria', 'Vencimento', 'Status', 'Observações'],
        'example': {'Tipo': 'pagar', 'Descrição': 'Aluguel', 'Valor': '1200.00',
                    'Categoria': 'Aluguel', 'Vencimento': '10/01/2025', 'Status': 'pendente', 'Observações': ''},
    },
    'clients': {
        'headers': ['Nome', 'Email', 'Telefone', 'Documento', 'Endereço', 'Observações'],
        'example': {'Nome': 'João da Silva', 'Email': 'joao@email.com', 'Telefone': '(44) 99999-0000',
                    'Documento': '123.456.789-00', 'Endereço': 'Rua das Flores, 100', 'Observações': ''},
    },
    'products': {
        'headers': ['Nome', 'SKU', 'Tipo', 'Categoria', 'Preço de Venda', 'Custo',
                    'Estoque', 'Estoque Mínimo', 'Unidade', 'Descrição'],
        'example': {'Nome': 'Produto Exemplo', 'SKU': 'PROD-001', 'Tipo': 'produto',
                    'Categoria': 'Geral', 'Preço de Venda': '50.00', 'Custo': '25.00',
                    'Estoque': '10', 'Estoque Mínimo': '2', 'Unidade': 'un', 'Descrição': ''},
    },
    'team': {
        'headers': ['Nome', 'Email', 'Role', 'Ativo'],
        'example': {'Nome': 'Maria Souza', 'Email': 'maria@empresa.com',
                    'Role': 'seller', 'Ativo': 'Sim'},
    },
}


@import_export_bp.route('/export/template/<module>', methods=['GET'])
@jwt_required()
def download_template(module):
    if module not in TEMPLATES:
        return jsonify({'error': 'Módulo inválido'}), 400
    tpl    = TEMPLATES[module]
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=tpl['headers'])
    writer.writeheader()
    writer.writerow(tpl['example'])
    output.seek(0)
    return Response(
        '\ufeff' + output.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': f'attachment; filename="template_{module}.csv"',
            'Access-Control-Expose-Headers': 'Content-Disposition',
        }
    )


# ─────────────────────────────────────────────
# EXPORTAR EQUIPE (somente admin)
# ─────────────────────────────────────────────

@import_export_bp.route('/export/team', methods=['GET'])
@jwt_required()
@admin_required
def export_team():
    user = get_current_user()
    fmt  = request.args.get('format', 'csv')

    members = User.query.filter_by(company_id=user.company_id).all()

    headers = ['ID', 'Nome', 'Email', 'Role', 'Tipo de Conta', 'Email Verificado', 'Ativo']
    rows = []
    for m in members:
        rows.append({
            'ID':               safe(m, 'id'),
            'Nome':             safe(m, 'name'),
            'Email':            safe(m, 'email'),
            'Role':             safe(m, 'role'),
            'Tipo de Conta':    safe(m, 'account_type'),
            'Email Verificado': 'Sim' if safe(m, 'email_verified', False) else 'Não',
            'Ativo':            'Sim' if safe(m, 'active', True) else 'Não',
        })

    stamp = datetime.now().strftime('%Y%m%d')
    return build_response(rows, headers, f'equipe_{stamp}', fmt)


# ─────────────────────────────────────────────
# IMPORTAÇÃO — PREVIEW
# ─────────────────────────────────────────────

@import_export_bp.route('/import/preview', methods=['POST'])
@jwt_required()
def import_preview():
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400

    file    = request.files['file']
    fname   = file.filename or ''
    columns = []
    rows    = []

    # ── xlsx ──
    if fname.endswith('.xlsx') or fname.endswith('.xls'):
        try:
            import openpyxl
            wb      = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
            ws      = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if all_rows:
                columns = [str(c) if c is not None else '' for c in all_rows[0]]
                for row in all_rows[1:6]:
                    rows.append({columns[i]: (str(row[i]) if row[i] is not None else '') for i in range(len(columns))})
        except ImportError:
            return jsonify({'error': 'openpyxl não instalado no servidor. Use CSV por enquanto.'}), 500
    else:
        # ── csv ──
        content = file.read().decode('utf-8-sig')
        reader  = csv.DictReader(io.StringIO(content))
        columns = list(reader.fieldnames or [])
        for i, row in enumerate(reader):
            if i >= 5:
                break
            rows.append(dict(row))

    # Detecta sistema de origem
    col_set = set(c.lower() for c in columns)
    detected = 'generico'
    if 'competência' in col_set or 'plano de contas' in col_set:
        detected = 'conta_azul'
    elif 'histórico' in col_set and 'débito' in col_set and 'crédito' in col_set:
        detected = 'nibo'
    elif 'serviço' in col_set and 'profissional' in col_set:
        detected = 'app_barber'

    return jsonify({
        'columns':         columns,
        'detected_system': detected,
        'preview_rows':    rows,
        'total_columns':   len(columns),
    })